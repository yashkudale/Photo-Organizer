# import modules
import os 
import sys
import glob
import tkinter as tk
from PIL import ImageTk, Image, ImageOps
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Print or not to print Debug Information ***************************
DEBUG = False

def log(s):
    if DEBUG:
        print(s)

# *******************************************************************

# Get Folder Information ********************************************

# Get current directory
script_directory = os.path.dirname(os.path.abspath(sys.argv[0])) 
print("\nDirectory Name: " + script_directory)

# Get the list of all files and directories
global img_files
global numOfFiles

img_files = []

for ext in ('*.gif', '*.png', '*.jpg'):
    img_files.extend(glob.glob(os.path.join(script_directory, ext).split("\\")[-1]))


numOfFiles = len(img_files)
print("\nTotal "+ str(numOfFiles) + " image files detected")

for i in range(numOfFiles):
    log(str(i)+ ': ' + img_files[i])

# Setup Output File
# Fixed Values
FILE_NAME = "OutputFile.xlsx"
WORK_SHEET_NAME = "Photo Rating"
BLACK = "000000"
THIN = Side(border_style="thin", color=BLACK)
NONE = Side(border_style="none")

csv_files = glob.glob(os.path.join(script_directory, "*.xlsx"))

# *******************************************************************


# Excel Handling Fuctions *******************************************

def setupOutputFile():
    wb = Workbook()
    # Renaming Sheet
    wb['Sheet'].title = 'Photo Rating'
    ws = wb[WORK_SHEET_NAME]

    # Creating Headings for the table
    HEADING = ['Sr. No.', 'File Name', 'Rating', 1, 2, 3, 4, 5]
    rowStart = 1
    colStart = 1
    for i in range(0, len(HEADING)):
        currentCell = ws.cell(rowStart, colStart + i, value = HEADING[i])
        currentCell.font = Font( size=14, bold= True)
        currentCell.alignment = Alignment(horizontal= "center", vertical= "center")

    # Filling in the list of files
    rowStart = 2
    colStart = 1

    for i in range(0, numOfFiles):
        # Fill serial numbers
        currentCell = ws.cell(rowStart + i, colStart, value = i + 1)
        currentCell.alignment = Alignment(horizontal= "center", vertical= "center")
        # Fill file names
        ws.cell(rowStart + i, colStart + 1, value = img_files[i].split("\\")[-1])
        # Fill in 0 ratings
        currentCell = ws.cell(rowStart + i, colStart + 2, value = 0)
        currentCell.alignment = Alignment(horizontal= "center", vertical= "center")

    # Iterate over all columns and adjust their widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(FILE_NAME)

def removeFill():
    wb = load_workbook(filename= FILE_NAME)
    ws = wb[WORK_SHEET_NAME]
    currentRating = ws.cell(row=1, column=1)
    
    # Get cell data and fill rating 
    rowStart = 2
    colStart = 1

    for i in range(0, numOfFiles):
        for j in range(0, 5):
            currentCell = ws.cell(rowStart + i, j + 4)
            currentCell.fill = PatternFill(fill_type = "none")
            currentCell.border = Border(top=NONE, left=NONE, right=NONE, bottom=NONE)
    
    wb.save(FILE_NAME)

def fillRatingColor():
    removeFill()
    wb = load_workbook(filename= FILE_NAME)
    ws = wb[WORK_SHEET_NAME]
    currentRating = ws.cell(row=1, column=1)
    
    # Get cell data and fill rating 
    rowStart = 2
    colStart = 1

    for i in range(0, numOfFiles):
        # Fill serial numbers
        currentRating = ws.cell(rowStart + i, colStart + 2)
        log(currentRating.value)

        if currentRating.value > 0 and currentRating.value < 6:
            for j in range(0, currentRating.value):
                currentCell = ws.cell(rowStart + i, j + 4)
                currentCell.fill = PatternFill(start_color="FFDB51", fill_type = "solid")
                currentCell.border = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)
    
    wb.save(FILE_NAME)

def setRating(imageNumber, rating):
    wb = load_workbook(filename= FILE_NAME)
    ws = wb[WORK_SHEET_NAME]
    
    # Set to rating column and row
    rowStart = 2 + imageNumber
    colStart = 3
    ws.cell(rowStart, colStart).value = rating
    wb.save(FILE_NAME)

def getRating(imageNumber):
    wb = load_workbook(filename= FILE_NAME)
    ws = wb[WORK_SHEET_NAME]
    
    # Set to rating column and row
    rowStart = 2 + imageNumber
    colStart = 3
    rating = ws.cell(rowStart, colStart).value
    return rating

def getFileName(imageNumber):
    wb = load_workbook(filename= FILE_NAME)
    ws = wb[WORK_SHEET_NAME]
    
    # Set to rating column and row
    rowStart = 2 + imageNumber
    colStart = 2
    fileName = ws.cell(rowStart, colStart).value
    return fileName

# Check if there is existing output file
if not csv_files:
    print("Output file didn't exist in this folder, creating new file ", FILE_NAME)
    setupOutputFile()

# loop over the list of csv files 
for f in csv_files:
    if f.split("\\")[-1] == FILE_NAME:
        # print the location and filename 
        print('\nOutput file exist') 
        print('Output File Name:', f.split("\\")[-1]) 
        
    else:
        print("Output file didn't exist in this folder, creating new file ", FILE_NAME)
        setupOutputFile()

# *******************************************************************

global imageLable
global currentImageNo
currentImageNo = 0

# Fuctions to update image
def updateImage():
    global currentImageNo
    global img_files

    # Feed image first
    imgFile = img_files[currentImageNo]
    log("Current File: " + str(currentImageNo) + ". " + imgFile)

    # Exception handler for unsupported file formats 
    try:
        displayedImage = Image.open(imgFile)
    except Image.UnidentifiedImageError:
        print("Not an image file")
        defaultImage = "Image_not_available.png"
        displayedImage = Image.open(defaultImage)

    # Limiting the max size for the image
    displayedImage = ImageOps.contain(displayedImage, (1100,690))

    # Passing the image to Lable 
    displayedImage = ImageTk.PhotoImage(displayedImage)
    imageLable.configure(image=displayedImage)
    imageLable.image= displayedImage

    # Update rating
    ratingVar.set(str(getRating(currentImageNo)))

    # Update file name
    currentFileVar.set(getFileName(currentImageNo))

# Button fuctions
def rightKeyEvent(event):
    nextButtonPress()

def leftKeyEvent(event):
    prevButtonPress()

def nextButtonPress():
    global currentImageNo
    global numOfFiles

    if currentImageNo < (numOfFiles - 1):
        currentImageNo = currentImageNo + 1
        updateImage()

def prevButtonPress():
    global currentImageNo

    if currentImageNo > 0:
        currentImageNo = currentImageNo - 1
        updateImage()

def radioButtonClicked(inputRating):
    setRating(currentImageNo, inputRating)

def closing_event():
    fillRatingColor()
    os.startfile(FILE_NAME)
    window.destroy()

# General setup
window = tk.Tk()
window.title("Photo Organizer")
window.geometry('1500x700')

# Left Frame
leftFrame = tk.LabelFrame(window, text = "List of Files", padx=5, pady=5)
leftFrame.pack(padx=10, expand=True, fill=tk.Y, side=tk.LEFT, anchor=tk.W)


# Current image file name variable
currentFileVar = tk.StringVar()
currentFileText = tk.Entry(leftFrame, textvariable=currentFileVar).pack(padx=5, pady=5,fill=tk.X)
currentFileVar.set(getFileName(currentImageNo))

fileList = tk.Text(leftFrame, width=40)
fileList.pack(padx=5, pady=5, expand=True, fill=tk.Y, side=tk.LEFT)

# Right Frame
global rightFrame
rightFrame = tk.LabelFrame(window, padx=5, pady=5, width=1900)
rightFrame.pack(padx=10, pady=8, expand=False, fill=tk.Y, side=tk.LEFT)

for items in img_files:
    fileList.insert(tk.END, items + '\n')

# Feed image first
currentImageNo = 0
imgFile = img_files[currentImageNo]

# Exception handler for unsupported file formats 
try:
    displayedImage = Image.open(imgFile)
except Image.UnidentifiedImageError:
    print("Not an image file")
    defaultImage = "Image_not_available.png"
    displayedImage = Image.open(defaultImage)

# Limiting the max size for the image
displayedImage = ImageOps.contain(displayedImage, (1100,690))

# Passing the image to Lable 
displayedImage = ImageTk.PhotoImage(displayedImage)
imageLable = tk.Label(rightFrame, image=displayedImage, width=1900)
imageLable.pack(padx=10, pady=10, expand=True, fill=tk.X, side=tk.TOP, anchor=tk.N)

# Frame with all the controls
controlsFrame = tk.Frame(rightFrame, padx=5, pady=5)
controlsFrame.pack(expand=True, fill=tk.X, side=tk.TOP, anchor=tk.S)

# Previous Button 
previousButton = tk.Button(controlsFrame, text='< Previous', font=("Courier", 14,"bold"), command=prevButtonPress).pack(expand=True, fill=tk.BOTH, side=tk.LEFT, anchor=tk.N)

# Rating Variable
ratingVar = tk.IntVar()
ratingVar.set(str(getRating(currentImageNo)))
# List of ratings
RATINGS = [
    ("*", 1),
    ("**", 2),
    ("***", 3),
    ("****", 4),
    ("*****", 5),
]
# Radio Buttons for Ratings
for Text, rating in RATINGS:
    tk.Radiobutton(controlsFrame, text=Text,font=("Courier", 24,"bold"), variable=ratingVar, value=rating, padx=10, command= lambda: radioButtonClicked(ratingVar.get())).pack(expand=True,side=tk.LEFT)

# Next Button
nextButton = tk.Button(controlsFrame, text='Next >',  font=("Courier", 14,"bold"), command=nextButtonPress).pack(expand=True, fill=tk.BOTH, side=tk.LEFT, anchor=tk.N)

window.bind("<Right>", rightKeyEvent)
window.bind("<Left>", leftKeyEvent)
window.protocol("WM_DELETE_WINDOW", closing_event) 
window.mainloop()